"""
Testes para core/xlsx_converter.py.
Fixtures criadas programaticamente via openpyxl e csv stdlib.
"""
import csv
from pathlib import Path

import openpyxl
import pytest

from core.xlsx_converter import _celula_str, planilha_to_md

# ── Fixtures ────────────────────────────────────────────────────────────────

@pytest.fixture
def xlsx_simples(tmp_path) -> Path:
    """XLSX com uma sheet e 3 linhas (header + 2 dados)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.append(["Nome", "Valor", "Ativo"])
    ws.append(["Alpha", 100, True])
    ws.append(["Beta", 200, False])
    path = tmp_path / "simples.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def xlsx_multi_sheet(tmp_path) -> Path:
    """XLSX com 2 sheets."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["A", "B"])
    ws1.append([1, 2])

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["X", "Y"])
    ws2.append([10, 20])

    path = tmp_path / "multi.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def xlsx_vazio(tmp_path) -> Path:
    """XLSX com sheet sem dados."""
    wb = openpyxl.Workbook()
    path = tmp_path / "vazio.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def csv_simples(tmp_path) -> Path:
    """CSV básico UTF-8."""
    path = tmp_path / "dados.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Coluna1", "Coluna2"])
        writer.writerow(["val1", "val2"])
        writer.writerow(["val3", "val4"])
    return path


@pytest.fixture
def csv_bom(tmp_path) -> Path:
    """CSV com BOM (Excel export no Windows)."""
    path = tmp_path / "bom.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Nome", "Preço"])
        writer.writerow(["Produto A", "9.99"])
    return path


@pytest.fixture
def csv_latin1(tmp_path) -> Path:
    """CSV com encoding Latin-1 (common em exports PT-BR)."""
    path = tmp_path / "latin1.csv"
    with open(path, "w", newline="", encoding="latin-1") as f:
        writer = csv.writer(f)
        writer.writerow(["Descrição", "Valor"])
        writer.writerow(["Ação", "100"])
    return path


# ── Testes XLSX ──────────────────────────────────────────────────────────────

def test_xlsx_header_e_dados(xlsx_simples):
    """Header e dados aparecem como tabela Markdown."""
    md = planilha_to_md(xlsx_simples)
    assert "| Nome | Valor | Ativo |" in md
    assert "| --- | --- | --- |" in md
    assert "Alpha" in md
    assert "Beta" in md


def test_xlsx_sheet_name_como_heading(xlsx_simples):
    """Nome da sheet aparece como ## heading."""
    md = planilha_to_md(xlsx_simples)
    assert "## Dados" in md


def test_xlsx_multi_sheet(xlsx_multi_sheet):
    """Múltiplas sheets geram múltiplos headings."""
    md = planilha_to_md(xlsx_multi_sheet)
    assert "## Sheet1" in md
    assert "## Sheet2" in md
    assert "| A | B |" in md
    assert "| X | Y |" in md


def test_xlsx_vazio_nao_falha(xlsx_vazio):
    """XLSX sem dados não levanta exceção."""
    md = planilha_to_md(xlsx_vazio)
    assert isinstance(md, str)


def test_xlsx_arquivo_inexistente(tmp_path):
    """FileNotFoundError para arquivo inexistente."""
    with pytest.raises(FileNotFoundError):
        planilha_to_md(tmp_path / "nao_existe.xlsx")


def test_xlsx_corrompido(tmp_path):
    """RuntimeError para XLSX inválido."""
    f = tmp_path / "corrompido.xlsx"
    f.write_bytes(b"nao_e_xlsx")
    with pytest.raises(RuntimeError):
        planilha_to_md(f)


# ── Testes CSV ───────────────────────────────────────────────────────────────

def test_csv_basico(csv_simples):
    """CSV básico → tabela Markdown com header e dados."""
    md = planilha_to_md(csv_simples)
    assert "| Coluna1 | Coluna2 |" in md
    assert "| --- | --- |" in md
    assert "val1" in md
    assert "val3" in md


def test_csv_bom(csv_bom):
    """CSV com BOM (Excel) é decodificado corretamente."""
    md = planilha_to_md(csv_bom)
    assert "Nome" in md  # BOM não deve aparecer no header
    assert "Produto A" in md


def test_csv_latin1(csv_latin1):
    """CSV Latin-1 com acentos PT-BR decodifica sem erro."""
    md = planilha_to_md(csv_latin1)
    assert "Descrição" in md or "Descri" in md  # aceita ambos (encoding best-effort)
    assert "Ação" in md or "A" in md


def test_csv_arquivo_inexistente(tmp_path):
    """FileNotFoundError para CSV inexistente."""
    with pytest.raises(FileNotFoundError):
        planilha_to_md(tmp_path / "nao_existe.csv")


def test_csv_extensao_invalida(tmp_path):
    """ValueError para extensão não suportada."""
    f = tmp_path / "arquivo.xls"
    f.write_bytes(b"xls_content")
    with pytest.raises(ValueError):
        planilha_to_md(f)


# ── Integridade do bloco de tabela ──────────────────────────────────────────

def _sem_linha_vazia_em_tabela(md: str) -> bool:
    """True se nenhuma linha vazia aparece entre duas linhas de tabela."""
    linhas = md.split("\n")
    for i in range(1, len(linhas) - 1):
        if linhas[i].strip():
            continue
        if linhas[i - 1].startswith("|") and linhas[i + 1].startswith("|"):
            return False
    return True


def test_xlsx_tabela_sem_linha_em_branco(xlsx_simples):
    """Linhas da tabela ficam contíguas — linha vazia quebra GFM/Obsidian."""
    md = planilha_to_md(xlsx_simples)
    assert _sem_linha_vazia_em_tabela(md), md


def test_xlsx_multi_sheet_tabela_sem_linha_em_branco(xlsx_multi_sheet):
    """Múltiplas abas: cada bloco íntegro, separação só entre abas."""
    md = planilha_to_md(xlsx_multi_sheet)
    assert _sem_linha_vazia_em_tabela(md), md
    assert "## Sheet1\n\n| A | B |\n| --- | --- |\n| 1 | 2 |" in md
    assert "| 1 | 2 |\n\n## Sheet2" in md


def test_csv_tabela_sem_linha_em_branco(csv_simples):
    """CSV mantém tabela contígua (referência de comportamento correto)."""
    md = planilha_to_md(csv_simples)
    assert _sem_linha_vazia_em_tabela(md), md


# ── Testes do helper _celula_str ─────────────────────────────────────────────

def test_celula_str_none():
    assert _celula_str(None) == ""


def test_celula_str_pipe_escapado():
    assert _celula_str("a|b") == "a\\|b"


def test_celula_str_quebra_linha():
    assert _celula_str("linha1\nlinha2") == "linha1 linha2"


def test_celula_str_numero():
    assert _celula_str(42) == "42"
