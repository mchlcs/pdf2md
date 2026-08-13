"""
Converte planilhas (.xlsx, .csv) em Markdown.

Estratégia:
- .xlsx → openpyxl: cada sheet vira tabela Markdown com nome como heading
- .csv  → stdlib csv: uma tabela Markdown (sem deps extras)

Nota: .xls (formato legado) não suportado — requer xlrd e o formato está
obsoleto desde Excel 2007. Usuários devem converter para .xlsx primeiro.
"""
import csv
from pathlib import Path
from typing import Any

from core.utils import _linha_tabela_md, _sanitizar_celula_md, _validar_existencia

# Encodings tentados em cascade para CSV — latin-1 nunca falha (mapeia 256 bytes).
_CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")


def planilha_to_md(path: Path) -> str:
    """
    Converte planilha (.xlsx ou .csv) em Markdown.

    Dispatcha para o conversor correto pela extensão.

    Args:
        path: Caminho para o arquivo de planilha.

    Returns:
        String Markdown com conteúdo tabular extraído.

    Raises:
        FileNotFoundError: Se path não existe.
        ValueError: Se extensão não é suportada.
    """
    _validar_existencia(path)

    sufixo = path.suffix.lower()
    if sufixo == ".xlsx":
        return _xlsx_para_md(path)
    elif sufixo == ".csv":
        return _csv_para_md(path)
    else:
        raise ValueError(f"Extensão não suportada por planilha_to_md: {sufixo}")


def _xlsx_para_md(path: Path) -> str:
    """Converte .xlsx em Markdown usando openpyxl."""
    import openpyxl  # lazy import — evita custo no startup

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(
            "Falha ao abrir XLSX — arquivo corrompido ou formato inválido"
        ) from exc

    # Cada aba vira um bloco "## Nome\n\n<tabela>"; blocos separados por "\n\n".
    # Linhas da tabela usam "\n" simples — linha em branco no meio quebra a
    # renderização em GFM/Obsidian.
    partes: list[str] = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # FIX 6: duas passadas em vez de list(iter_rows(values_only=True))
            # — read_only=True é streaming; materializar 1M linhas em memória
            # derrotava o propósito. A 1ª passada acha a primeira linha
            # NÃO-VAZIA como header — antes, sheet com linhas vazias no topo
            # tinha "headers" todos vazios e o conteúdo era descartado
            # silenciosamente (só o heading `## nome` era emitido).
            header = _primeira_linha_nao_vazia(ws)
            if header is None:
                continue  # sheet sem nenhuma célula com conteúdo
            headers = [_celula_str(c) for c in header]

            linhas: list[str] = []
            linhas.append(_linha_tabela_md(headers))
            linhas.append(_linha_tabela_md(["---"] * len(headers)))

            # 2ª passada: corpo — pula a primeira linha não-vazia (o header
            # já usado acima) e linhas totalmente vazias.
            header_visto = False
            for row in ws.iter_rows(values_only=True):
                celulas = [_celula_str(c) for c in row]
                if not any(celulas):
                    continue
                if not header_visto:
                    header_visto = True
                    continue
                linhas.append(_linha_tabela_md(celulas))

            partes.append(f"## {sheet_name}\n\n" + "\n".join(linhas))
    finally:
        wb.close()

    return "\n\n".join(partes)


def _primeira_linha_nao_vazia(ws: Any) -> tuple[Any, ...] | None:
    """Primeira linha com ao menos uma célula com conteúdo (header real).

    FIX 6: em vez de assumir que a linha 0 é o header, varre até achar a
    primeira linha utilizável — planilhas com linhas em branco no topo
    (comuns em exportações) não perdem mais o cabeçalho nem o corpo.
    """
    for row in ws.iter_rows(values_only=True):
        if any(_celula_str(c) for c in row):
            return tuple(row)  # tuple(Any) → tuple[Any, ...] (mypy no-any-return)
    return None


def _csv_para_md(path: Path) -> str:
    """Converte .csv em Markdown usando stdlib csv. Lê bytes uma vez, decode in-memory."""
    dados = path.read_bytes()

    rows = None
    for enc in _CSV_ENCODINGS:
        try:
            texto = dados.decode(enc)
            rows = list(csv.reader(texto.splitlines()))
            break
        except UnicodeDecodeError:
            continue

    if not rows:
        return ""

    headers = [_celula_str(c) for c in rows[0]]
    linhas: list[str] = []
    linhas.append("| " + " | ".join(headers) + " |")
    linhas.append("| " + " | ".join(["---"] * len(headers)) + " |")

    for row in rows[1:]:
        celulas = [_celula_str(c) for c in row]
        linhas.append("| " + " | ".join(celulas) + " |")

    return "\n".join(linhas)


def _celula_str(valor: Any) -> str:
    """Normaliza valor de célula para string Markdown segura."""
    if valor is None:
        return ""
    return _sanitizar_celula_md(str(valor))
